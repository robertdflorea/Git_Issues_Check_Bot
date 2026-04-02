"""
GitHub Issue Finder - GUI Application
Finds valid closed GitHub issues with base SHA for AI coding tasks.
"""

import json
import os
from dotenv import load_dotenv
import re
import threading
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import webbrowser
import customtkinter as ctk
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

GITHUB_API      = "https://api.github.com"
VALID_LANGUAGES = ["Python", "JavaScript", "TypeScript"]
TEST_INDICATORS = ["test", "tests", "spec", "specs", "__tests__", "pytest.ini", "jest.config"]
CONFIG_PATH     = os.path.join(os.path.expanduser("~"), ".issue_finder.json")

# Load .env from the same folder as the exe / script
_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
load_dotenv(_env_path)

DEFAULT_REQUIREMENTS = """\
• Language must be Python, JavaScript, or TypeScript
• Repository size must be ≤ 200 MB
• Repository must have a test suite
• Repository must have a README file
• Repository must use a standard package manager (pip/npm/conda)
• Repository must not be archived or disabled
• Issue must be non-trivial (meaningful body, ≥ 2 comments)
• Issue must have a linked merged PR with a resolvable base SHA
• Issue title must not indicate a trivial change (typo, rename, etc.)"""

# ── Colours ────────────────────────────────────────────────────────────────────
PASS_COL = "#2ecc71"
FAIL_COL = "#e74c3c"
INFO_COL = "#3498db"
WARN_COL = "#f39c12"

DARK_THEME = {
    "BG_DARK":    "#12181f",
    "BG_MID":     "#1a2535",
    "BG_LEFT":    "#0f151c",
    "BG_LOG":     "#0d1117",
    "BG_EVEN":    "#1e2d40",
    "BG_CHECKED": "#1a3a1a",
    "FG":         "white",
    "FG_DIM":     "gray",
    "HDR_BG":     "#1F4E79",
    "HDR_FG":     "white",
    "SEL_BG":     "#2980b9",
    "LINE_COL":   "#2c3e50",
}

LIGHT_THEME = {
    "BG_DARK":    "#f0f4f8",
    "BG_MID":     "#e2e8f0",
    "BG_LEFT":    "#dde3ea",
    "BG_LOG":     "#f8fafc",
    "BG_EVEN":    "#cbd5e1",
    "BG_CHECKED": "#bbf7d0",
    "FG":         "#1a202c",
    "FG_DIM":     "#64748b",
    "HDR_BG":     "#1e40af",
    "HDR_FG":     "white",
    "SEL_BG":     "#3b82f6",
    "LINE_COL":   "#94a3b8",
}

# Active theme (mutable dict, updated on toggle)
T = dict(DARK_THEME)

# Convenience aliases kept for code that references bare names
def _t(key): return T[key]

BG_DARK  = T["BG_DARK"]
BG_MID   = T["BG_MID"]
BG_LEFT  = T["BG_LEFT"]
BG_LOG   = T["BG_LOG"]

FONT_BODY  = ("Segoe UI", 12)
FONT_BOLD  = ("Segoe UI", 12, "bold")
FONT_SMALL = ("Segoe UI", 11)
FONT_MONO  = ("Consolas", 11)
FONT_TITLE = ("Segoe UI", 15, "bold")
FONT_HEAD  = ("Segoe UI", 13, "bold")
FONT_LOG   = ("Consolas", 11)


# ── Config persistence ─────────────────────────────────────────────────────────

def load_config():
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_config(data):
    try:
        existing = load_config()
        existing.update(data)
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(existing, f)
    except Exception:
        pass


# ── URL parsing ────────────────────────────────────────────────────────────────

def parse_input_url(url):
    """
    Returns:
      ("issue", "owner/repo", issue_number)  — for issue URLs
      ("repo",  "owner/repo")                — for repo URLs
      None                                   — invalid
    """
    url = url.strip().rstrip("/")
    m = re.search(r"github\.com[/:]([^/]+)/([^/]+)/issues/(\d+)$", url)
    if m:
        return ("issue", f"{m.group(1)}/{m.group(2)}", int(m.group(3)))
    m = re.search(r"github\.com[/:]([^/]+/[^/]+?)(?:\.git)?$", url)
    if m:
        return ("repo", m.group(1))
    if re.fullmatch(r"[^/]+/[^/]+", url):
        return ("repo", url)
    return None


def parse_requirements(req_text):
    t = req_text.lower()
    return {
        "language":      any(w in t for w in ["language", "python", "javascript", "typescript"]),
        "size":          any(w in t for w in ["size", "200", "mb"]),
        "tests":         "test" in t,
        "readme":        "readme" in t,
        "pkg_manager":   any(w in t for w in ["package manager", "pip", "npm", "conda"]),
        "archived":      any(w in t for w in ["archive", "disabled"]),
        "nontrivial":    "non-trivial" in t or "nontrivial" in t,
        "linked_pr":     any(w in t for w in ["linked", "pr", "pull request", "base sha"]),
        "trivial_title": any(w in t for w in ["typo", "rename", "trivial"]),
    }


# ── GitHub API helpers ─────────────────────────────────────────────────────────

def get_headers(token=None):
    h = {"Accept": "application/vnd.github+json", "X-GitHub-Api-Version": "2022-11-28"}
    if token:
        h["Authorization"] = f"Bearer {token}"
    return h


def api_get(url, headers, params=None, retry=3, stop_event=None):
    for attempt in range(retry):
        if stop_event and stop_event.is_set():
            return None
        try:
            r = requests.get(url, headers=headers, params=params, timeout=12)
            if r.status_code == 403 and "rate limit" in r.text.lower():
                wait = int(r.headers.get("Retry-After", 30))
                for _ in range(wait):
                    if stop_event and stop_event.is_set():
                        return None
                    time.sleep(1)
                continue
            return r
        except requests.RequestException:
            if attempt == retry - 1:
                return None
            time.sleep(1)


def validate_repo(headers, full_name, checks, stop_event=None):
    r = api_get(f"{GITHUB_API}/repos/{full_name}", headers, stop_event=stop_event)
    if not r or r.status_code != 200:
        return None, f"Could not fetch repo"
    repo = r.json()
    if checks["language"] and repo.get("language", "") not in VALID_LANGUAGES:
        return None, f"Language '{repo.get('language','')}' not Python/JS/TS"
    if checks["size"] and repo.get("size", 0) / 1024 > 200:
        return None, f"Repo too large ({repo.get('size',0)/1024:.1f} MB)"
    if checks["archived"]:
        if repo.get("archived"):
            return None, "Repo is archived"
        if repo.get("disabled"):
            return None, "Repo is disabled"
    return repo, None


def check_repo_contents(headers, full_name, branch, stop_event=None):
    r = api_get(f"{GITHUB_API}/repos/{full_name}/git/trees/{branch}",
                headers, stop_event=stop_event)
    if not r or r.status_code != 200:
        return {"tests": False, "readme": False, "pkg": False, "dockerfile": False}
    paths = [item["path"].lower() for item in r.json().get("tree", [])]
    return {
        "tests":      any(any(ind in p for ind in TEST_INDICATORS) for p in paths),
        "readme":     any("readme" in p for p in paths),
        "pkg":        any(p in paths for p in [
                          "package.json", "setup.py", "requirements.txt",
                          "pyproject.toml", "pipfile", "environment.yml"]),
        "dockerfile": any("dockerfile" in p for p in paths),
    }


def fetch_closed_issues(headers, repo_full, stop_event=None):
    issues, page = [], 1
    while True:
        if stop_event and stop_event.is_set():
            break
        r = api_get(f"{GITHUB_API}/repos/{repo_full}/issues", headers,
                    {"state": "closed", "per_page": 100, "page": page,
                     "sort": "updated", "direction": "desc"},
                    stop_event=stop_event)
        if not r or r.status_code != 200:
            break
        batch = [i for i in r.json() if "pull_request" not in i]
        issues.extend(batch)
        if len(r.json()) < 100 or page >= 5:
            break
        page += 1
    return issues


def fetch_pr_files(headers, repo_full, pr_number, stop_event=None):
    """Return (file_paths, total_changed_lines) for a PR (up to 100 files)."""
    r = api_get(f"{GITHUB_API}/repos/{repo_full}/pulls/{pr_number}/files",
                headers, {"per_page": 100}, stop_event=stop_event)
    if not r or r.status_code != 200:
        return [], 0
    files = r.json()
    paths = [f["filename"] for f in files]
    lines = sum(f.get("additions", 0) + f.get("deletions", 0) for f in files)
    return paths, lines


def find_linked_pr(headers, repo_full, issue_number, stop_event=None):
    """Returns (pr_url, base_sha, changed_files_count, changed_file_list, changed_lines)."""
    t_headers = {**headers, "Accept": "application/vnd.github.mockingbird-preview+json"}
    t_resp = api_get(f"{GITHUB_API}/repos/{repo_full}/issues/{issue_number}/timeline",
                     t_headers, stop_event=stop_event)
    pr_numbers = []
    if t_resp and t_resp.status_code == 200:
        for event in t_resp.json():
            if event.get("event") in ("cross-referenced", "closed"):
                issue_data = event.get("source", {}).get("issue", {})
                if issue_data.get("pull_request"):
                    num = issue_data.get("number")
                    if num:
                        pr_numbers.append(num)
    if not pr_numbers:
        if stop_event and stop_event.is_set():
            return None, None, 0, [], 0
        q = f"repo:{repo_full} is:pr is:closed {issue_number}"
        sr = api_get(f"{GITHUB_API}/search/issues", headers,
                     {"q": q, "per_page": 10}, stop_event=stop_event)
        if sr and sr.status_code == 200:
            for item in sr.json().get("items", []):
                body = (item.get("body") or "").lower()
                refs = [f"#{issue_number}", f"fixes #{issue_number}",
                        f"closes #{issue_number}", f"resolves #{issue_number}"]
                if any(ref in body for ref in refs):
                    pr_numbers.append(item["number"])
    seen = set()
    for num in pr_numbers:
        if stop_event and stop_event.is_set():
            return None, None, 0, [], 0
        if num in seen:
            continue
        seen.add(num)
        pr_r = api_get(f"{GITHUB_API}/repos/{repo_full}/pulls/{num}",
                       headers, stop_event=stop_event)
        if pr_r and pr_r.status_code == 200:
            pr = pr_r.json()
            sha = pr.get("base", {}).get("sha", "")
            if sha:
                changed_count = pr.get("changed_files", 0)
                file_list, changed_lines = fetch_pr_files(headers, repo_full, num, stop_event)
                return pr.get("html_url", ""), sha, changed_count, file_list, changed_lines
    return None, None, 0, [], 0


def check_issue_nontrivial(issue, checks):
    results = []
    title    = issue.get("title", "").lower()
    body     = issue.get("body") or ""
    comments = issue.get("comments", 0)
    if checks["trivial_title"]:
        trivial_kw = [
            "typo", "rename", "spelling", "grammar", "bump version", "update readme",
            "update changelog", "update docs", "update documentation", "fix typo",
            "fix spelling", "fix grammar", "css only", "css fix", "doc only",
            "docs only", "documentation only", "move file", "move files",
            "clean up", "cleanup", "formatting", "whitespace", "lint fix",
        ]
        results.append(("Title not trivial", not any(kw in title for kw in trivial_kw)))
    if checks["nontrivial"]:
        results.append(("Body length ≥ 50 chars", len(body) >= 50))
        results.append(("Comments ≥ 2",           comments >= 2))
    return results


def validate_single_issue(headers, repo_full, issue_number, repo, contents,
                           checks, log_cb, stop_event):
    """Validate one issue and return a result dict or None."""
    r = api_get(f"{GITHUB_API}/repos/{repo_full}/issues/{issue_number}",
                headers, stop_event=stop_event)
    if not r or r.status_code != 200:
        log_cb("error", f"  ✗ Could not fetch issue #{issue_number}")
        return None
    issue = r.json()
    if "pull_request" in issue:
        log_cb("error", "  ✗ That URL points to a Pull Request, not an Issue")
        return None
    return _validate_issue_obj(headers, repo_full, issue, repo, contents,
                               checks, log_cb, stop_event)


def _validate_issue_obj(headers, repo_full, issue, repo, contents,
                        checks, log_cb, stop_event):
    num   = issue["number"]
    title = issue["title"]
    lang    = repo.get("language", "")
    size_mb = round(repo.get("size", 0) / 1024, 1)

    log_cb("issue_start", f"#{num}: {title[:70]}")

    issue_checks = check_issue_nontrivial(issue, checks)

    for label, ok in issue_checks:
        log_cb("check", (label, ok))

    issue_ok = all(ok for _, ok in issue_checks)

    pr_url, base_sha, changed_files, changed_file_list, changed_lines = None, None, 0, [], 0
    if issue_ok:
        log_cb("info", "  → Looking for linked PR…")
        pr_url, base_sha, changed_files, changed_file_list, changed_lines = find_linked_pr(
            headers, repo_full, num, stop_event)
        pr_ok = bool(base_sha)
        issue_checks.append(("Linked PR with base SHA", pr_ok))
        log_cb("check", ("Linked PR with base SHA", pr_ok))
        issue_ok = pr_ok

        if issue_ok:
            CODE_EXTS = {".py", ".js", ".ts", ".tsx", ".jsx"}
            code_files = [
                f for f in changed_file_list
                if os.path.splitext(f)[1].lower() in CODE_EXTS
            ]
            code_file_count = len(code_files)
            enough_code_files = code_file_count > 2
            issue_checks.append((f"Code files changed > 2 (found {code_file_count})", enough_code_files))
            log_cb("check", (f"Code files changed > 2 (found {code_file_count})", enough_code_files))
            issue_ok = enough_code_files

            # Also reject pure test-only or doc-only PRs
            if issue_ok:
                non_test_code = [
                    f for f in code_files
                    if not any(ind in f.lower() for ind in TEST_INDICATORS)
                ]
                has_non_test_code = len(non_test_code) > 0
                issue_checks.append(("PR touches non-test code files", has_non_test_code))
                log_cb("check", ("PR touches non-test code files", has_non_test_code))
                issue_ok = has_non_test_code

        if issue_ok:
            enough_lines = changed_lines > 20
            issue_checks.append((f"Meaningful lines changed > 20 (found {changed_lines})", enough_lines))
            log_cb("check", (f"Meaningful lines changed > 20 (found {changed_lines})", enough_lines))
            issue_ok = enough_lines
    else:
        # still record the PR check as skipped / not checked
        issue_checks.append(("Linked PR with base SHA", False))
        log_cb("check", ("Linked PR with base SHA", False))

    if issue_ok:
        log_cb("pass", f"  ✓ Issue #{num} is VALID")
        return {
            "repo":          repo_full,
            "language":      lang,
            "size_mb":       size_mb,
            "issue_number":  num,
            "issue_title":   title,
            "issue_url":     issue["html_url"],
            "pr_url":        pr_url or "",
            "base_sha":      base_sha,
            "changed_files":      changed_files,
            "changed_file_list":  changed_file_list,
            "changed_lines":      changed_lines,
            "dockerfile":         contents.get("dockerfile", False),
            "checks":        issue_checks,
        }
    else:
        log_cb("fail", f"  ✗ Issue #{num} failed validation")
        return None


def run_search(token, input_type, repo_full, issue_number,
               req_text, log_cb, add_result_cb, done_cb, stop_event):
    headers = get_headers(token)
    checks  = parse_requirements(req_text)

    # ── Repo validation ────────────────────────────────────────────────
    log_cb("section", "Validating Repository")
    repo, err = validate_repo(headers, repo_full, checks, stop_event)
    if stop_event.is_set():
        done_cb(); return
    if err:
        log_cb("error", f"✗ {err}")
        done_cb(); return

    lang    = repo.get("language", "")
    size_mb = round(repo.get("size", 0) / 1024, 1)
    branch  = repo.get("default_branch", "main")

    repo_checks = [
        (f"Language: {lang}",         lang in VALID_LANGUAGES or not checks["language"]),
        (f"Size: {size_mb} MB ≤ 200", size_mb <= 200          or not checks["size"]),
        ("Not archived / disabled",    not repo.get("archived") and not repo.get("disabled")),
    ]
    contents = check_repo_contents(headers, repo_full, branch, stop_event)
    if stop_event.is_set():
        done_cb(); return

    if checks["tests"]:
        repo_checks.append(("Has test suite",           contents["tests"]))
    if checks["readme"]:
        repo_checks.append(("Has README",               contents["readme"]))
    if checks["pkg_manager"]:
        repo_checks.append(("Has package manager file", contents["pkg"]))

    log_cb("repo_checks", repo_checks)

    failed_repo = [lbl for lbl, ok in repo_checks if not ok]
    if failed_repo:
        log_cb("error", "✗ Repo failed: " + ", ".join(failed_repo))
        done_cb(); return

    # ── Single issue mode ──────────────────────────────────────────────
    if input_type == "issue":
        log_cb("section", f"Validating Issue #{issue_number}")
        result = validate_single_issue(headers, repo_full, issue_number,
                                       repo, contents, checks, log_cb, stop_event)
        if result:
            add_result_cb(result)
        done_cb()
        return

    # ── Repo mode: fetch all closed issues ─────────────────────────────
    log_cb("section", "Fetching Closed Issues")
    raw_issues = fetch_closed_issues(headers, repo_full, stop_event)
    if stop_event.is_set():
        done_cb(); return
    log_cb("info", f"Found {len(raw_issues)} closed issues — validating each…")
    log_cb("section", "Validating Issues")

    for issue in raw_issues:
        if stop_event.is_set():
            break
        result = _validate_issue_obj(headers, repo_full, issue, repo, contents,
                                     checks, log_cb, stop_event)
        if result:
            add_result_cb(result)
        time.sleep(0.12)

    done_cb()


# ── Excel export ───────────────────────────────────────────────────────────────

def _files_cell(r):
    """Format changed files as single-line: count | lines | filenames."""
    file_list = r.get("changed_file_list", [])
    count     = r.get("changed_files", len(file_list))
    lines     = r.get("changed_lines", 0)
    names     = ", ".join(os.path.basename(f) for f in file_list) if file_list else "—"
    return f"{count} files | {lines} lines | {names}"


def export_to_excel(results, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Valid Issues"

    hfill   = PatternFill("solid", fgColor="1F4E79")
    hfont   = Font(bold=True, color="FFFFFF", size=12)
    altfill = PatternFill("solid", fgColor="D6E4F0")
    bs      = Side(style="thin", color="AAAAAA")
    bdr     = Border(left=bs, right=bs, top=bs, bottom=bs)

    headers    = ["Repo Name", "Issue Link", "Issue Title", "Base SHA",
                  "Repo Category", "PR Link", "Files Changed", "Status"]
    col_widths = [40, 55, 55, 45, 16, 55, 45, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 24

    for i, r in enumerate(results, 1):
        fill = altfill if i % 2 == 0 else None
        for col, val in enumerate([
            r["repo"], r["issue_url"], r["issue_title"], r["base_sha"],
            r["language"], r["pr_url"],
            _files_cell(r), "Open",
        ], 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            if fill: cell.fill = fill
            cell.border = bdr
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (1,2,3,4,6,7)))
        n_files = len(r.get("changed_file_list", []))
        ws.row_dimensions[i+1].height = max(20, 15 * (n_files + 1))

    wb.save(path)


# ── Spinner ────────────────────────────────────────────────────────────────────

class Spinner(ctk.CTkFrame):
    FRAMES = ["⠋","⠙","⠹","⠸","⠼","⠴","⠦","⠧","⠇","⠏"]

    def __init__(self, parent, **kwargs):
        super().__init__(parent, fg_color="transparent", **kwargs)
        self._frame = 0; self._job = None; self._active = False
        self._spin = ctk.CTkLabel(self, text="",
                                  font=ctk.CTkFont(family="Segoe UI", size=16),
                                  text_color=INFO_COL, width=22)
        self._spin.pack(side="left")
        self._text = ctk.CTkLabel(self, text="",
                                  font=ctk.CTkFont(*FONT_SMALL), text_color="gray")
        self._text.pack(side="left", padx=(4, 0))

    def start(self, text="Searching…"):
        self._active = True
        self._text.configure(text=text)
        self._tick()

    def stop(self, text=""):
        self._active = False
        if self._job: self.after_cancel(self._job); self._job = None
        self._spin.configure(text="")
        self._text.configure(text=text)

    def set_text(self, text):
        self._text.configure(text=text[:60])

    def _tick(self):
        if not self._active: return
        self._spin.configure(text=self.FRAMES[self._frame % len(self.FRAMES)])
        self._frame += 1
        self._job = self.after(90, self._tick)


# ── Main application ───────────────────────────────────────────────────────────

# Treeview column definitions
COLUMNS = [
    ("Check",         "✓",              42,  "center"),
    ("Repo",          "Repository",     200, "w"),
    ("Issue URL",     "Issue Link",     260, "w"),
    ("Title",         "Issue Title",    320, "w"),
    ("Base SHA",      "Base SHA",       180, "w"),
    ("Category",      "Repo Category",  130, "center"),
    ("PR URL",        "PR Link",        240, "w"),
    ("Changed Files", "Files Changed",  420, "w"),
]

URL_COLS = {"Issue URL", "PR URL"}   # double-click opens browser


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("GitHub Issue Finder")
        self.geometry("1420x860")
        self.minsize(1000, 640)
        self.results     = []
        self._stop_event = threading.Event()
        self._sel_col    = None
        self._sort_col   = None
        self._sort_rev   = False
        self._checked    = set()   # set of treeview item iids that are ticked
        self._load_saved_theme()
        self._build_ui()
        self._load_saved_token()

    # ── Theme / token persistence ──────────────────────────────────────────────

    def _load_saved_theme(self):
        cfg   = load_config()
        theme = cfg.get("theme", "dark")
        if theme == "light":
            T.update(LIGHT_THEME)
            ctk.set_appearance_mode("light")
        else:
            T.update(DARK_THEME)
            ctk.set_appearance_mode("dark")

    # ──────────────────────────────────────────────────────────────────────────

    def _load_saved_token(self):
        # Priority: 1) saved config  2) .env file  3) empty
        cfg   = load_config()
        token = cfg.get("token", "") or os.getenv("GITHUB_TOKEN", "")
        if token:
            self.token_entry.insert(0, token)

    def _save_token(self):
        save_config({"token": self.token_entry.get().strip()})

    # ── UI ─────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Dark.TPanedwindow", background=BG_DARK)

        self._h_pane = ttk.PanedWindow(self, orient="horizontal",
                                       style="Dark.TPanedwindow")
        self._h_pane.pack(fill="both", expand=True)

        # ── LEFT ──────────────────────────────────────────────────────────
        _left_host = tk.Frame(self._h_pane, bg=T["BG_LEFT"])
        self._h_pane.add(_left_host, weight=0)

        self.left = ctk.CTkScrollableFrame(_left_host, fg_color=BG_LEFT, corner_radius=0)
        self.left.pack(fill="both", expand=True)

        # Make left panel contents fill horizontally
        self.left.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(self.left, text="GitHub Issue Finder",
                     font=ctk.CTkFont(*FONT_TITLE)).pack(pady=(22, 2), fill="x", padx=18)
        ctk.CTkLabel(self.left, text="Find valid closed issues with base SHA",
                     font=ctk.CTkFont(*FONT_SMALL), text_color="gray"
                     ).pack(pady=(0, 18), fill="x", padx=18)

        def slabel(text):
            ctk.CTkLabel(self.left, text=text, font=ctk.CTkFont(*FONT_BOLD),
                         anchor="w").pack(anchor="w", padx=18, pady=(8, 1), fill="x")

        def hlabel(text):
            ctk.CTkLabel(self.left, text=text, font=ctk.CTkFont(*FONT_SMALL),
                         text_color="#6c7a89", anchor="w").pack(anchor="w", padx=18, fill="x")

        # Repo / Issue URL
        slabel("Repository or Issue URL")
        hlabel("Repo: github.com/owner/repo")
        hlabel("Issue: github.com/owner/repo/issues/123")
        self.repo_entry = ctk.CTkEntry(
            self.left, placeholder_text="https://github.com/owner/repo",
            height=36, font=ctk.CTkFont(*FONT_BODY))
        self.repo_entry.pack(padx=18, pady=(4, 14), fill="x")

        # Token
        slabel("GitHub Token  (optional)")
        hlabel("Saved automatically between sessions")
        self.token_entry = ctk.CTkEntry(
            self.left, placeholder_text="ghp_…",
            show="*", height=36, font=ctk.CTkFont(*FONT_BODY))
        self.token_entry.pack(padx=18, pady=(4, 14), fill="x")

        # Requirements
        slabel("Validation Requirements")
        hlabel("Edit to enable / disable each check")
        self.req_box = ctk.CTkTextbox(
            self.left, height=200,
            font=ctk.CTkFont(*FONT_SMALL), wrap="word")
        self.req_box.pack(padx=18, pady=(4, 18), fill="x")
        self.req_box.insert("1.0", DEFAULT_REQUIREMENTS)

        # Buttons — all fill="x" so they resize with panel
        self.search_btn = ctk.CTkButton(
            self.left, text="  Find Issues", height=42,
            font=ctk.CTkFont(*FONT_BOLD), command=self._start_search)
        self.search_btn.pack(padx=18, pady=(0, 8), fill="x")

        self.stop_btn = ctk.CTkButton(
            self.left, text="  Stop & Show Results", height=38,
            fg_color="#c0392b", hover_color="#922b21",
            font=ctk.CTkFont(*FONT_BOLD), command=self._stop_search, state="disabled")
        self.stop_btn.pack(padx=18, pady=(0, 8), fill="x")

        self.export_btn = ctk.CTkButton(
            self.left, text="  Export to Excel", height=38,
            fg_color="#27ae60", hover_color="#1e8449",
            font=ctk.CTkFont(*FONT_BOLD), command=self._export_excel, state="disabled")
        self.export_btn.pack(padx=18, pady=(0, 8), fill="x")

        self.clear_btn = ctk.CTkButton(
            self.left, text="  Clear", height=36,
            fg_color="gray35", hover_color="gray25",
            font=ctk.CTkFont(*FONT_BOLD), command=self._clear)
        self.clear_btn.pack(padx=18, pady=(0, 8), fill="x")

        # Row height slider
        slabel("Row Height")
        self._row_height_var = tk.IntVar(value=30)
        _rh_row = ctk.CTkFrame(self.left, fg_color="transparent")
        _rh_row.pack(padx=18, pady=(2, 8), fill="x")
        self._rh_slider = ctk.CTkSlider(
            _rh_row, from_=22, to=120,
            variable=self._row_height_var,
            command=self._on_row_height_change)
        self._rh_slider.pack(side="left", fill="x", expand=True)
        self._rh_lbl = ctk.CTkLabel(
            _rh_row, text="30 px", width=52,
            font=ctk.CTkFont(*FONT_SMALL), anchor="e")
        self._rh_lbl.pack(side="left", padx=(6, 0))

        self.theme_btn = ctk.CTkButton(
            self.left, text="☀  Switch to Light Theme", height=36,
            fg_color="#4a5568", hover_color="#2d3748",
            font=ctk.CTkFont(*FONT_BOLD), command=self._toggle_theme)
        self.theme_btn.pack(padx=18, pady=(0, 8), fill="x")

        self.spinner = Spinner(self.left)
        self.spinner.pack(padx=18, pady=(12, 4), fill="x")

        self.stat_lbl = ctk.CTkLabel(self.left, text="",
                                     font=ctk.CTkFont(*FONT_SMALL), text_color=PASS_COL,
                                     anchor="w")
        self.stat_lbl.pack(padx=18, pady=(0, 18), fill="x")

        # ── RIGHT: vertical paned window ──────────────────────────────────
        _right_host = tk.Frame(self._h_pane, bg=T["BG_DARK"])
        self._h_pane.add(_right_host, weight=1)
        self._right_host = _right_host

        self._v_pane = ttk.PanedWindow(_right_host, orient="vertical",
                                       style="Dark.TPanedwindow")
        self._v_pane.pack(fill="both", expand=True, padx=8, pady=8)

        # ── Results table (top) ────────────────────────────────────────
        _tbl_host = tk.Frame(self._v_pane, bg=T["BG_DARK"])
        self._v_pane.add(_tbl_host, weight=3)
        self._tbl_host = _tbl_host

        tbl_wrap = ctk.CTkFrame(_tbl_host, fg_color=BG_DARK, corner_radius=0)
        tbl_wrap.pack(fill="both", expand=True)

        tbl_hdr = ctk.CTkFrame(tbl_wrap, fg_color="transparent")
        tbl_hdr.pack(fill="x", padx=12, pady=(10, 4))
        ctk.CTkLabel(tbl_hdr, text="Results  (click column header to sort)",
                     font=ctk.CTkFont(*FONT_HEAD)).pack(side="left")
        self._res_count_lbl = ctk.CTkLabel(tbl_hdr, text="",
                                           font=ctk.CTkFont(*FONT_SMALL), text_color="gray")
        self._res_count_lbl.pack(side="left", padx=10)

        self._progress = ctk.CTkProgressBar(tbl_wrap, mode="indeterminate", height=4)
        self._progress.pack(fill="x", padx=12, pady=(0, 4))
        self._progress.set(0)

        tree_host = ctk.CTkFrame(tbl_wrap, fg_color=BG_MID, corner_radius=7)
        tree_host.pack(fill="both", expand=True, padx=10, pady=(0, 8))

        self._tree_style = style
        self._apply_treeview_style()

        col_ids = [c[0] for c in COLUMNS]
        self.tree = ttk.Treeview(tree_host, columns=col_ids, show="headings",
                                 style="Issue.Treeview", selectmode="browse")

        for col_id, col_label, col_w, col_anchor in COLUMNS:
            self.tree.column(col_id, width=col_w, anchor=col_anchor, minwidth=40)
            if col_id == "Check":
                # Header click = select all / deselect all
                self.tree.heading(col_id, text=col_label,
                                  command=self._toggle_all_checks)
            else:
                self.tree.heading(col_id, text=col_label,
                                  command=lambda c=col_id: self._sort_by(c))

        sy = ttk.Scrollbar(tree_host, orient="vertical",   command=self.tree.yview)
        sx = ttk.Scrollbar(tree_host, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        sy.pack(side="right",  fill="y")
        sx.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)
        self._tree_host = tree_host   # kept for bg update on theme toggle
        self._refresh_tree_tags()

        # ── Vertical column lines on body (1-px Frame separators) ────────
        # Use thin tk.Frame strips instead of an opaque canvas overlay so
        # that the treeview rows remain visible and clickable.
        self._sep_frames = []
        n_seps = len(COLUMNS) - 1
        for _ in range(n_seps):
            f = tk.Frame(tree_host, bg=T["LINE_COL"], width=1, bd=0,
                         highlightthickness=0)
            self._sep_frames.append(f)

        def _draw_vlines(event=None):
            self.tree.update_idletasks()
            tx   = self.tree.winfo_x()
            ty   = self.tree.winfo_y()
            th   = self.tree.winfo_height()
            head_h = 30
            try:
                head_h = int(self.tree.tk.call(
                    "ttk::treeview::HeaderHeight", self.tree))
            except Exception:
                pass
            x = tx
            body_top  = ty + head_h
            body_h    = max(1, th - head_h)
            for i, (col_id, _, _, _) in enumerate(COLUMNS[:-1]):
                try:
                    x += self.tree.column(col_id, "width")
                except Exception:
                    break
                self._sep_frames[i].place(
                    x=x - 1, y=body_top, width=1, height=body_h)
                self._sep_frames[i].lift()   # keep above tree but below scrollbars

        self.tree.bind("<Configure>",       _draw_vlines)
        self.tree.bind("<B1-Motion>",       _draw_vlines)   # live column drag
        self.tree.bind("<ButtonRelease-1>", _draw_vlines)   # after column drag release
        self._draw_vlines = _draw_vlines
        self.after(400, _draw_vlines)

        col_ids_list = [c[0] for c in COLUMNS]

        def _get_col_id(event):
            col_num = self.tree.identify_column(event.x)
            if not col_num:
                return None
            idx = int(col_num.lstrip("#")) - 1
            return col_ids_list[idx] if 0 <= idx < len(col_ids_list) else None

        def _on_click(event):
            region = self.tree.identify_region(event.x, event.y)
            if region != "cell":
                return
            col_id = _get_col_id(event)
            self._sel_col = self.tree.identify_column(event.x)
            iid = self.tree.identify_row(event.y)
            if not iid:
                return
            # Toggle checkbox column
            if col_id == "Check":
                if iid in self._checked:
                    self._checked.discard(iid)
                else:
                    self._checked.add(iid)
                self._refresh_row_display(iid)
                self._update_export_label()

        def _on_double_click(event):
            region = self.tree.identify_region(event.x, event.y)
            if region != "cell":
                return
            col_id = _get_col_id(event)
            iid    = self.tree.identify_row(event.y)
            if not iid or col_id not in URL_COLS:
                return
            idx = col_ids_list.index(col_id)
            vals = self.tree.item(iid, "values")
            url  = str(vals[idx]) if idx < len(vals) else ""
            if url.startswith("http"):
                webbrowser.open(url)

        self.tree.bind("<ButtonRelease-1>", _on_click)
        self.tree.bind("<Double-ButtonRelease-1>", _on_double_click)

        def _copy_cell(event=None):
            row = self.tree.focus()
            if not row or not self._sel_col: return
            idx = int(self._sel_col.lstrip("#")) - 1
            vals = self.tree.item(row, "values")
            if 0 <= idx < len(vals):
                self.clipboard_clear(); self.clipboard_append(str(vals[idx]))

        self.tree.bind("<Control-c>", _copy_cell)

        self._ctx = tk.Menu(self, tearoff=0, bg="#1a2535", fg="white",
                            activebackground="#2980b9", activeforeground="white",
                            font=FONT_SMALL, bd=0, relief="flat")

        def _copy_row():
            row = self.tree.focus()
            if not row: return
            self.clipboard_clear()
            self.clipboard_append("\t".join(str(v) for v in self.tree.item(row, "values")))

        def _copy_named(name):
            row = self.tree.focus()
            if not row: return
            col_ids_list = [c[0] for c in COLUMNS]
            if name in col_ids_list:
                idx = col_ids_list.index(name)
                vals = self.tree.item(row, "values")
                if idx < len(vals):
                    self.clipboard_clear(); self.clipboard_append(str(vals[idx]))

        self._ctx.add_command(label="  Copy Cell",      command=_copy_cell)
        self._ctx.add_command(label="  Copy Row",       command=_copy_row)
        self._ctx.add_separator()
        self._ctx.add_command(label="  Copy Issue Link", command=lambda: _copy_named("Issue URL"))
        self._ctx.add_command(label="  Copy Base SHA",   command=lambda: _copy_named("Base SHA"))
        self._ctx.add_command(label="  Copy PR Link",    command=lambda: _copy_named("PR URL"))

        def _show_ctx(event):
            iid = self.tree.identify_row(event.y)
            if iid:
                self.tree.selection_set(iid)
                self.tree.focus(iid)
                self._sel_col = self.tree.identify_column(event.x)
            self._ctx.post(event.x_root, event.y_root)

        self.tree.bind("<Button-3>", _show_ctx)

        # ── Cell editor (click to select/edit/copy) ────────────────────
        self._cell_editor = None

        self.tree.bind("<Double-ButtonRelease-1>", _on_double_click)  # keep URL open
        self.tree.bind("<ButtonRelease-1>", _on_click)               # keep checkbox
        self.tree.bind("<Control-Return>", lambda e: None)
        # Single click (non-checkbox) opens inline editor on second click
        self._last_clicked_cell = (None, None)

        def _on_click_editor(event):
            _on_click(event)   # still do checkbox logic
            region = self.tree.identify_region(event.x, event.y)
            if region != "cell":
                return
            col_id = _get_col_id(event)
            iid    = self.tree.identify_row(event.y)
            if col_id and col_id not in ("Check",) and iid:
                if self._last_clicked_cell == (iid, col_id):
                    self._show_cell_editor(iid, col_id)
                self._last_clicked_cell = (iid, col_id)

        # Replace the earlier ButtonRelease binding
        self.tree.unbind("<ButtonRelease-1>")
        self.tree.bind("<ButtonRelease-1>", _on_click_editor)

        # ── Live log panel (bottom) ────────────────────────────────────
        _log_host = tk.Frame(self._v_pane, bg=T["BG_DARK"])
        self._v_pane.add(_log_host, weight=1)
        self._log_host = _log_host

        log_wrap = ctk.CTkFrame(_log_host, fg_color=BG_DARK, corner_radius=0)
        log_wrap.pack(fill="both", expand=True)

        log_hdr = ctk.CTkFrame(log_wrap, fg_color="transparent")
        log_hdr.pack(fill="x", padx=12, pady=(8, 2))
        ctk.CTkLabel(log_hdr, text="Validation Log",
                     font=ctk.CTkFont(*FONT_HEAD)).pack(side="left")
        self._log_count_lbl = ctk.CTkLabel(log_hdr, text="",
                                           font=ctk.CTkFont(*FONT_SMALL), text_color="gray")
        self._log_count_lbl.pack(side="left", padx=10)

        self._log_box = ctk.CTkTextbox(
            log_wrap, fg_color=BG_LOG, text_color="white",
            font=ctk.CTkFont(*FONT_LOG), wrap="word", state="disabled")
        self._log_box.pack(fill="both", expand=True, padx=10, pady=(0, 8))

        # colour tags via underlying tk Text widget
        self._log_box._textbox.tag_configure("section", foreground="#9b59b6",
                                              font=("Consolas", 11, "bold"))
        self._log_box._textbox.tag_configure("pass",    foreground=PASS_COL)
        self._log_box._textbox.tag_configure("fail",    foreground=FAIL_COL)
        self._log_box._textbox.tag_configure("info",    foreground=INFO_COL)
        self._log_box._textbox.tag_configure("warn",    foreground=WARN_COL)
        self._log_box._textbox.tag_configure("check_pass", foreground=PASS_COL)
        self._log_box._textbox.tag_configure("check_fail", foreground=FAIL_COL,
                                              font=("Consolas", 11, "bold"))
        self._log_box._textbox.tag_configure("issue",   foreground="#ecf0f1",
                                              font=("Consolas", 11, "bold"))

        self._issue_count = 0

        # Initial sash
        self.after(300, lambda: self._v_pane.sashpos(0, 520))

        # Sync theme button label with loaded theme
        if ctk.get_appearance_mode().lower() == "light":
            self.theme_btn.configure(text="🌙  Switch to Dark Theme")
        else:
            self.theme_btn.configure(text="☀  Switch to Light Theme")

    # ── Sorting ────────────────────────────────────────────────────────────────

    # ── Theme ──────────────────────────────────────────────────────────────────

    def _apply_treeview_style(self):
        s = self._tree_style
        rh = getattr(self, "_row_height_var", None)
        rh = rh.get() if rh else 30
        s.configure("Issue.Treeview",
                    background=T["BG_MID"], foreground=T["FG"],
                    fieldbackground=T["BG_MID"], rowheight=rh,
                    font=FONT_MONO, borderwidth=0)
        s.configure("Issue.Treeview.Heading",
                    background=T["HDR_BG"], foreground=T["HDR_FG"],
                    font=("Segoe UI", 12, "bold"), relief="raised", padding=(6, 4))
        s.map("Issue.Treeview",
              background=[("selected", T["SEL_BG"])],
              foreground=[("selected", "white")])
        # Draw vertical lines by configuring the separator
        s.configure("Issue.TSeparator", background=T["LINE_COL"])

    def _refresh_tree_tags(self):
        self.tree.tag_configure("odd",     background=T["BG_MID"])
        self.tree.tag_configure("even",    background=T["BG_EVEN"])
        self.tree.tag_configure("checked", background=T["BG_CHECKED"])

    def _toggle_theme(self):
        T.clear()
        if ctk.get_appearance_mode().lower() == "dark":
            T.update(LIGHT_THEME)
            ctk.set_appearance_mode("light")
            self.theme_btn.configure(text="🌙  Switch to Dark Theme")
            save_config({"theme": "light"})
        else:
            T.update(DARK_THEME)
            ctk.set_appearance_mode("dark")
            self.theme_btn.configure(text="☀  Switch to Light Theme")
            save_config({"theme": "dark"})

        self._apply_treeview_style()
        self._refresh_tree_tags()
        # Re-apply tags to existing rows
        for iid in self.tree.get_children():
            if iid in self._checked:
                self.tree.item(iid, tags=("checked",))
            else:
                idx = self.tree.index(iid)
                self.tree.item(iid, tags=("odd" if (idx + 1) % 2 else "even",))

        # Update all plain tk.Frame backgrounds
        for frame, key in [
            (self._right_host, "BG_DARK"),
            (self._tbl_host,   "BG_DARK"),
            (self._log_host,   "BG_DARK"),
        ]:
            try: frame.configure(bg=T[key])
            except Exception: pass

        # Update paned window sash color
        try:
            self._tree_style.configure("Dark.TPanedwindow", background=T["BG_DARK"])
        except Exception:
            pass

        # Update log box colors
        try:
            self._log_box.configure(fg_color=T["BG_LOG"], text_color=T["FG"])
        except Exception:
            pass

        # Update tree host (the CTkFrame wrapping the treeview)
        try:
            self._tree_host.configure(fg_color=T["BG_MID"])
        except Exception:
            pass

        # Redraw vertical lines with new color
        try:
            for f in self._sep_frames:
                f.configure(bg=T["LINE_COL"])
            self._draw_vlines()
        except Exception:
            pass

        # Update context menu colors
        try:
            self._ctx.configure(bg=T["BG_MID"], fg=T["FG"],
                                 activebackground=T["SEL_BG"])
        except Exception:
            pass

    # ── Row height ─────────────────────────────────────────────────────────────

    def _on_row_height_change(self, value):
        h = int(float(value))
        self._rh_lbl.configure(text=f"{h} px")
        self._apply_treeview_style()
        # Force the treeview to redraw at the new row height
        self.tree.update_idletasks()
        try:
            self._draw_vlines()
        except Exception:
            pass

    # ── Select all / deselect all ──────────────────────────────────────────────

    def _toggle_all_checks(self):
        all_iids = self.tree.get_children()
        if not all_iids:
            return
        # If all are checked → deselect all, otherwise select all
        if all(iid in self._checked for iid in all_iids):
            for iid in all_iids:
                self._checked.discard(iid)
        else:
            for iid in all_iids:
                self._checked.add(iid)
        for iid in all_iids:
            self._refresh_row_display(iid)
        self._update_export_label()

    # ── Sort ───────────────────────────────────────────────────────────────────

    def _sort_by(self, col_id):
        if self._sort_col == col_id:
            self._sort_rev = not self._sort_rev
        else:
            self._sort_col = col_id
            self._sort_rev = False

        # Update heading arrows (skip Check column)
        for c_id, c_label, _, _ in COLUMNS:
            if c_id == "Check":
                continue
            arrow = ""
            if c_id == self._sort_col:
                arrow = "  ▲" if not self._sort_rev else "  ▼"
            self.tree.heading(c_id, text=c_label + arrow)

        key_map = {
            "Repo":          lambda r: r["repo"].lower(),
            "Issue URL":     lambda r: r["issue_url"],
            "Title":         lambda r: r["issue_title"].lower(),
            "Base SHA":      lambda r: r["base_sha"],
            "Category":      lambda r: r["language"],
            "PR URL":        lambda r: r["pr_url"],
            "Changed Files": lambda r: r.get("changed_files", 0),
        }
        key_fn = key_map.get(col_id, lambda r: "")
        self.results.sort(key=key_fn, reverse=self._sort_rev)
        self._repopulate_table()

    def _repopulate_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for i, r in enumerate(self.results, 1):
            iid = r.get("_iid")
            tag = "checked" if iid in self._checked else ("odd" if i % 2 else "even")
            new_iid = self.tree.insert("", "end", tags=(tag,), values=self._row_values(r, iid))
            r["_iid"] = new_iid

    def _row_values(self, r, iid=None):
        check     = "☑" if iid and iid in self._checked else "☐"
        file_list = r.get("changed_file_list", [])
        count     = r.get("changed_files", len(file_list))
        lines     = r.get("changed_lines", 0)
        names     = ", ".join(os.path.basename(f) for f in file_list) if file_list else "—"
        files_cell = f"{count} files | {lines} lines | {names}"
        return (
            check,
            r["repo"],
            r["issue_url"],
            r["issue_title"],
            r["base_sha"],
            r["language"],
            r["pr_url"],
            files_cell,
        )

    def _refresh_row_display(self, iid):
        """Update checkbox cell and row background for a single row."""
        idx = self.tree.index(iid)
        if 0 <= idx < len(self.results):
            vals = list(self.tree.item(iid, "values"))
            vals[0] = "☑" if iid in self._checked else "☐"
            self.tree.item(iid, values=tuple(vals))
        # Background: checked overrides odd/even
        if iid in self._checked:
            self.tree.item(iid, tags=("checked",))
        else:
            tag = "odd" if (self.tree.index(iid) + 1) % 2 else "even"
            self.tree.item(iid, tags=(tag,))

    def _update_export_label(self):
        n = len(self._checked)
        if n:
            self.export_btn.configure(
                text=f"  Export {n} Ticked Row{'s' if n != 1 else ''} to Excel")
        else:
            self.export_btn.configure(text="  Export to Excel")

    def _show_cell_editor(self, iid, col_id):
        """Open a floating, selectable Text editor over the clicked cell."""
        if self._cell_editor:
            try:
                self._cell_editor.destroy()
            except Exception:
                pass
            self._cell_editor = None

        bbox = self.tree.bbox(iid, col_id)
        if not bbox:
            return
        x, y, w, h = bbox

        # Fetch cell value
        col_ids_list = [c[0] for c in COLUMNS]
        if col_id not in col_ids_list:
            return
        idx  = col_ids_list.index(col_id)
        vals = self.tree.item(iid, "values")
        text = str(vals[idx]) if idx < len(vals) else ""

        # Count lines to determine editor height
        lines     = text.split("\n")
        line_h    = 18
        ed_height = max(h, min(len(lines) * line_h + 8, 300))
        ed_width  = max(w, 260)

        editor = tk.Text(
            self.tree,
            width=1, height=1,
            font=FONT_MONO,
            bg=T["BG_MID"], fg=T["FG"],
            insertbackground=T["FG"],
            selectbackground=T["SEL_BG"], selectforeground="white",
            relief="solid", borderwidth=1,
            wrap="none",
        )
        editor.place(x=x, y=y, width=ed_width, height=ed_height)
        editor.insert("1.0", text)
        editor.focus_set()
        editor.tag_add("sel", "1.0", "end")

        def _close(event=None):
            # Save edited value back to treeview
            new_val = editor.get("1.0", "end-1c")
            vals_list = list(self.tree.item(iid, "values"))
            if idx < len(vals_list):
                vals_list[idx] = new_val
                self.tree.item(iid, values=tuple(vals_list))
                # Also sync back to results
                tree_idx = self.tree.index(iid)
                if 0 <= tree_idx < len(self.results):
                    result = self.results[tree_idx]
                    field_map = {
                        "Repo":          "repo",
                        "Title":         "issue_title",
                        "Issue URL":     "issue_url",
                        "Base SHA":      "base_sha",
                        "Category":      "language",
                        "PR URL":        "pr_url",
                    }
                    if col_id in field_map:
                        result[field_map[col_id]] = new_val
            try:
                editor.destroy()
            except Exception:
                pass
            self._cell_editor = None

        editor.bind("<Escape>", _close)
        editor.bind("<FocusOut>", _close)
        self._cell_editor = editor

    # ── Log helpers ────────────────────────────────────────────────────────────

    def _log_write(self, text, tag=""):
        self._log_box.configure(state="normal")
        tb = self._log_box._textbox
        if tag:
            tb.insert("end", text + "\n", tag)
        else:
            tb.insert("end", text + "\n")
        tb.see("end")
        self._log_box.configure(state="disabled")

    def _on_log(self, kind, data):
        if kind == "section":
            self._log_write(f"\n── {data} ──", "section")

        elif kind == "info":
            self._log_write(f"  {data}", "info")
            self.spinner.set_text(str(data))

        elif kind == "repo_checks":
            for label, ok in data:
                tag = "check_pass" if ok else "check_fail"
                icon = "✓" if ok else "✗"
                self._log_write(f"  {icon}  {label}", tag)

        elif kind == "error":
            self._log_write(f"  {data}", "fail")
            self.spinner.set_text(str(data)[:60])

        elif kind == "issue_start":
            self._issue_count += 1
            self._log_write(f"\n  [{self._issue_count}] {data}", "issue")
            self._log_count_lbl.configure(
                text=f"({self._issue_count} issues checked)")

        elif kind == "check":
            label, ok = data
            tag = "check_pass" if ok else "check_fail"
            icon = "✓" if ok else "✗"
            self._log_write(f"      {icon}  {label}", tag)

        elif kind == "pass":
            self._log_write(f"  {data}", "pass")

        elif kind == "fail":
            self._log_write(f"  {data}", "fail")

    def _log_cb(self, kind, data):
        self.after(0, lambda k=kind, d=data: self._on_log(k, d))

    # ── Callbacks ──────────────────────────────────────────────────────────────

    def _add_result_cb(self, result):
        self.after(0, lambda r=result: self._append_row(r))

    def _append_row(self, r):
        self.results.append(r)
        i   = len(self.results)
        tag = "odd" if i % 2 else "even"
        iid = self.tree.insert("", "end", tags=(tag,), values=self._row_values(r))
        r["_iid"] = iid
        self.tree.yview_moveto(1.0)
        self._res_count_lbl.configure(
            text=f"({i} valid issue{'s' if i != 1 else ''} found)", text_color=PASS_COL)
        self.export_btn.configure(state="normal")

    # ── Actions ────────────────────────────────────────────────────────────────

    def _start_search(self):
        raw = self.repo_entry.get().strip()
        parsed = parse_input_url(raw)
        if not parsed:
            messagebox.showerror("Invalid URL",
                                 "Please enter a valid GitHub repository or issue URL.\n"
                                 "Repo:  https://github.com/owner/repo\n"
                                 "Issue: https://github.com/owner/repo/issues/123")
            return

        token    = self.token_entry.get().strip() or None
        req_text = self.req_box.get("1.0", "end")

        # Save token
        self._save_token()

        input_type   = parsed[0]
        repo_full    = parsed[1]
        issue_number = parsed[2] if input_type == "issue" else None

        self._stop_event.clear()
        self.results = []
        self._issue_count = 0
        self.search_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self.export_btn.configure(state="disabled")
        self._clear_table()
        self._clear_log()
        self.stat_lbl.configure(text="")
        label = f"Issue #{issue_number}" if input_type == "issue" else repo_full
        self.spinner.start(f"Searching {label}…")
        self._progress.start()

        def run():
            run_search(token, input_type, repo_full, issue_number, req_text,
                       log_cb=self._log_cb,
                       add_result_cb=self._add_result_cb,
                       done_cb=lambda: self.after(0, self._on_done),
                       stop_event=self._stop_event)

        threading.Thread(target=run, daemon=True).start()

    def _stop_search(self):
        self._stop_event.set()
        self.stop_btn.configure(state="disabled")
        self.spinner.set_text("Stopping…")

    def _on_done(self):
        self._progress.stop()
        self._progress.set(0)
        self.search_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled")
        n = len(self.results)
        if n:
            self.spinner.stop(f"Done — {n} valid issue(s) found")
            self.stat_lbl.configure(text=f"✓  {n} valid issue(s) found", text_color=PASS_COL)
            self.export_btn.configure(state="normal")
        else:
            self.spinner.stop("No valid issues found")
            self.stat_lbl.configure(text="✗  No valid issues found", text_color=FAIL_COL)

    def _clear_table(self):
        for item in self.tree.get_children(): self.tree.delete(item)
        self._checked.clear()
        self._res_count_lbl.configure(text="")
        self.export_btn.configure(text="  Export to Excel")
        for c_id, c_label, _, _ in COLUMNS:
            self.tree.heading(c_id, text=c_label)
        self._sort_col = None; self._sort_rev = False

    def _clear_log(self):
        self._log_box.configure(state="normal")
        self._log_box._textbox.delete("1.0", "end")
        self._log_box.configure(state="disabled")
        self._log_count_lbl.configure(text="")

    def _clear(self):
        self.results = []
        self._clear_table()
        self._clear_log()
        self.export_btn.configure(state="disabled")
        self.stat_lbl.configure(text="")
        self.spinner.stop("")

    def _export_excel(self):
        if not self.results: return
        # Export only ticked rows if any are ticked, otherwise all
        if self._checked:
            to_export = [r for r in self.results if r.get("_iid") in self._checked]
        else:
            to_export = self.results
        if not to_export: return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="valid_issues.xlsx")
        if not path: return
        try:
            export_to_excel(to_export, path)
            n = len(to_export)
            messagebox.showinfo("Exported",
                                f"Saved {n} row{'s' if n != 1 else ''} to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export failed", str(e))


if __name__ == "__main__":
    app = App()
    app.mainloop()
